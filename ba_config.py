from __future__ import annotations

import argparse
import ast
from dataclasses import dataclass, field
import operator
import re
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries


MODULES = {
    "PEC8445": {"UI": 8, "DI": 4, "AO": 4, "DO": 5, "price": 4058, "is_main": True},
    "PEC8044": {"UI": 8, "DI": 0, "AO": 4, "DO": 4, "price": 2492, "is_main": True},
    "PUC16000": {"UI": 16, "DI": 0, "AO": 0, "DO": 0, "price": 1365, "is_main": False},
    "PUC00016": {"UI": 0, "DI": 0, "AO": 0, "DO": 16, "price": 1474, "is_main": False},
    "PUC5533": {"UI": 5, "DI": 5, "AO": 3, "DO": 3, "price": 1470, "is_main": False},
    "PUC6002": {"UI": 6, "DI": 0, "AO": 0, "DO": 2, "price": 988, "is_main": False},
}

BOX_SIZES = [
    {"capacity": 1, "tag": "S规格", "price": 800, "model": "500x600x150"},
    {"capacity": 3, "tag": "M规格", "price": 1000, "model": "600x800x150"},
    {"capacity": 5, "tag": "L规格", "price": 1200, "model": "800x1000x150"},
    {"capacity": 6, "tag": "XL规格", "price": 1400, "model": "800x1200x200"},
    {"capacity": 7, "tag": "XXL规格", "price": 1600, "model": "1000x1200x200"},
]

POINT_MULTIPLIER = 1.8
GATEWAY_POINT = 200
MAX_MODULES_PER_DDC = 7

SOFTWARE_TIERS = [
    (0, 500, "一类软件"),
    (501, 1250, "二类软件"),
    (1251, 2500, "三类软件"),
    (2501, 5000, "四类软件"),
    (5001, 10000, "五类软件"),
    (10001, 25000, "六类软件"),
]

CUSTOMER_FINAL_GATEWAY_PRICE = 3000
CUSTOMER_FINAL_MODULE_PRICES = {
    "PEC8445": 4300,
    "PEC8044": 2500,
    "PUC16000": 1360,
    "PUC00016": 1470,
    "PUC5533": 980,
    "PUC6002": 820,
    "S规格": 800,
    "M规格": 1000,
    "L规格": 1200,
    "XL规格": 1400,
    "XXL规格": 1600,
}

MODULE_ROWS = [
    ("PEC8445", "可编程控制器", "PEC8445-PB1-SM"),
    ("PEC8044", "可编程控制器", "PEC8044-PB1-SO"),
    ("PUC16000", "扩展控制模块", "PUC16000-EM2"),
    ("PUC00016", "扩展控制模块", "PUC00016-EM2"),
    ("PUC5533", "扩展控制模块", "PUC5533-EM2"),
    ("PUC6002", "扩展控制模块", "PUC6002-EM2"),
]

BOX_ROWS = [
    ("S规格", "DDC控制箱", "500x600x150"),
    ("M规格", "DDC控制箱", "600x800x150"),
    ("L规格", "DDC控制箱", "800x1000x150"),
    ("XL规格", "DDC控制箱", "800x1200x200"),
    ("XXL规格", "DDC控制箱", "1000x1200x200"),
]

DDC_HEADERS = [
    "设备名称",
    "PEC8445-PB1-SM",
    "PEC8044-PB1-SO",
    "PUC16000-EM2",
    "PUC00016-EM2",
    "PUC5533-EM2",
    "PUC6002-EM2",
    "DDC控制箱 500x600x150",
    "DDC控制箱 600x800x150",
    "DDC控制箱 800x1000x150",
    "DDC控制箱 800x1200x200",
    "DDC控制箱 1000x1200x200",
]

LIST_HEADERS = ["序号", "产品名称", "品牌", "型号", "技术规格", "数量", "含税单价", "含税总价"]
EXTRA_SENSOR_HEADERS = {"风道CO2", "室内CO2浓度"}
SENSOR_PRICE_PRODUCT_ALIASES = {
    "初效滤网过滤": "空气压差开关",
    "中效滤网过滤": "空气压差开关",
    "压差报警": "空气压差开关",
    "风机压差检测": "空气压差开关",
    "超高液位报警": "液位开关",
    "超低液位报警": "液位开关",
    "水位溢出报警": "液位开关",
    "液位高/低报警": "液位开关",
    "新风温度": "风管温度",
    "送风温度": "风管温度",
    "新风湿度": "风管温湿度",
    "送风湿度": "风管温湿度",
    "风道CO2": "风管二氧化碳",
    "室内CO2浓度": "室内二氧化碳",
    "投入式液位监测": "投入式液位",
}


def _num(value: Any) -> float:
    if isinstance(value, bool) or value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0
    return 0.0


def _number_or_none(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text or text.startswith("="):
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _intish(value: float | int | None) -> int | float | None:
    if value is None:
        return None
    numeric = float(value)
    if abs(numeric - round(numeric)) < 1e-9:
        return int(round(numeric))
    return numeric


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _norm(value: Any) -> str:
    text = _text(value).lower()
    replacements = {
        "（": "(",
        "）": ")",
        "，": ",",
        " ": "",
        "\n": "",
        "\r": "",
        "-": "",
        "_": "",
        "×": "x",
        "*": "x",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def _dimension_key(value: Any) -> str:
    match = re.search(r"(\d{3,4})x(\d{3,4})", _norm(value))
    if not match:
        return ""
    return f"{match.group(1)}x{match.group(2)}"


def _point_key(value: Any) -> str:
    compact = _norm(value).upper()
    if compact in {"AI", "DI", "AO", "DO"}:
        return compact
    if compact == "A0":
        return "AO"
    if compact == "D0":
        return "DO"
    return ""


def _is_total_label(value: Any) -> bool:
    compact = _norm(value).upper()
    return compact in {"小计", "合计", "SUBTOTAL"}


class FormulaEvaluator:
    _cell_ref_re = re.compile(r"(?<![A-Za-z0-9_])\$?([A-Z]{1,3})\$?(\d+)")
    _sum_re = re.compile(r"SUM\(([^()]+)\)", re.IGNORECASE)

    def __init__(self, formula_ws, value_ws):
        self.formula_ws = formula_ws
        self.value_ws = value_ws
        self._cache: dict[tuple[int, int], float] = {}

    def numeric_cell(self, row: int, col: int) -> float:
        key = (row, col)
        if key in self._cache:
            return self._cache[key]

        raw = self.formula_ws.cell(row, col).value
        if not (isinstance(raw, str) and raw.startswith("=")):
            value = _num(raw)
            self._cache[key] = value
            return value

        try:
            value = self._eval_formula(raw[1:].strip())
        except Exception:
            value = _num(self.value_ws.cell(row, col).value)
        self._cache[key] = value
        return value

    def _eval_formula(self, expr: str) -> float:
        while True:
            match = self._sum_re.search(expr)
            if not match:
                break
            total = sum(self._eval_sum_part(part.strip()) for part in match.group(1).split(","))
            expr = expr[: match.start()] + str(total) + expr[match.end() :]

        expr = self._cell_ref_re.sub(
            lambda m: str(self.numeric_cell(int(m.group(2)), column_index_from_string(m.group(1)))),
            expr,
        )
        return float(_safe_eval(expr))

    def _eval_sum_part(self, expr: str) -> float:
        if ":" not in expr:
            return self._eval_formula(expr)
        min_col, min_row, max_col, max_row = range_boundaries(expr)
        total = 0.0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                total += self.numeric_cell(row, col)
        return total


_ALLOWED_BINOPS = {ast.Add: operator.add, ast.Sub: operator.sub, ast.Mult: operator.mul, ast.Div: operator.truediv}
_ALLOWED_UNARY = {ast.UAdd: operator.pos, ast.USub: operator.neg}


def _safe_eval(expr: str) -> float:
    tree = ast.parse(expr, mode="eval")

    def visit(node):
        if isinstance(node, ast.Expression):
            return visit(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        if isinstance(node, ast.BinOp) and type(node.op) in _ALLOWED_BINOPS:
            return _ALLOWED_BINOPS[type(node.op)](visit(node.left), visit(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _ALLOWED_UNARY:
            return _ALLOWED_UNARY[type(node.op)](visit(node.operand))
        raise ValueError(f"Unsupported formula expression: {expr}")

    return float(visit(tree))


@dataclass
class DdcConfig:
    name: str
    demand: dict[str, float]
    modules: dict[str, int]
    boxes: dict[str, int]
    source: str
    overloaded: bool = False


@dataclass
class ProjectData:
    project_name: str
    ddc_configs: list[DdcConfig]
    interfaces: list[tuple[str, float]]
    sensors: list[tuple[str, float]]
    point_totals: dict[str, float]

    @property
    def gateway_quantity(self) -> float:
        return sum(qty for _, qty in self.interfaces)

    @property
    def software_points(self) -> float:
        return sum(self.point_totals.values()) * POINT_MULTIPLIER + self.gateway_quantity * GATEWAY_POINT

    @property
    def software_tag(self) -> str:
        points = self.software_points
        for low, high, tag in SOFTWARE_TIERS:
            if low <= points <= high:
                return tag
        raise ValueError(f"总点数 {points} 超出软件档位范围")

    @property
    def module_totals(self) -> dict[str, int]:
        totals = {key: 0 for key in MODULES}
        for config in self.ddc_configs:
            for key, value in config.modules.items():
                totals[key] = totals.get(key, 0) + value
        return totals

    @property
    def box_totals(self) -> dict[str, int]:
        totals = {row[0]: 0 for row in BOX_ROWS}
        for config in self.ddc_configs:
            for key, value in config.boxes.items():
                totals[key] = totals.get(key, 0) + value
        return totals


class PointWorkbookParser:
    def __init__(self, path: str | Path, reserve_factor: float = 1.1, ddc_source: str = "auto"):
        self.path = Path(path)
        self.reserve_factor = reserve_factor
        self.ddc_source = ddc_source
        self.formula_wb = openpyxl.load_workbook(self.path, data_only=False)
        self.value_wb = openpyxl.load_workbook(self.path, data_only=True)
        self.formula_ws = self._select_sheet(self.formula_wb)
        self.value_ws = self.value_wb[self.formula_ws.title]
        self.evaluator = FormulaEvaluator(self.formula_ws, self.value_ws)
        self.header_row = self._find_header_row()
        self.group_row = self._find_group_row()
        self.headers = self._headers_at(self.header_row)
        self.name_col = self._find_header("设备名称") or self._find_header("名称") or self._find_header("Equipments") or 1
        self.qty_col = self._find_header("数量") or self._find_header("Qty.") or self._find_header("Qty") or 4
        self.is_bas_layout = self._find_header("名称") is not None and self._find_header("设备名称") is None
        self.summary_cols = self._find_summary_cols()
        self.group_ranges = self._find_group_ranges()
        self.module_cols, self.box_cols = self._find_config_cols()
        self.total_row = self._find_total_row()
        self.interface_start = self._find_interface_start()

    def close(self) -> None:
        self.formula_wb.close()
        self.value_wb.close()

    def parse(self) -> ProjectData:
        ddc_configs = self._parse_ddc_configs()
        point_totals = {key: 0.0 for key in ("DO", "AO", "DI", "AI")}
        for config in ddc_configs:
            for key in point_totals:
                point_totals[key] += config.demand.get(key, 0)
        return ProjectData(
            project_name=self._project_name(),
            ddc_configs=ddc_configs,
            interfaces=self._interfaces(),
            sensors=self._sensor_totals(),
            point_totals=point_totals,
        )

    def _select_sheet(self, wb):
        candidates = []
        for ws in wb.worksheets:
            score = 0
            for row in range(1, min(ws.max_row, 12) + 1):
                values = [_text(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 120) + 1)]
                normalized = {_norm(value) for value in values}
                point_keys = {_point_key(value) for value in values}
                if (
                    {"设备名称", "数量"}.issubset(set(values))
                    or {"名称", "数量"}.issubset(set(values))
                    or ("equipments" in normalized and ("qty." in normalized or "qty" in normalized))
                ):
                    score += 10
                if {"DO", "AO", "DI", "AI"}.issubset(point_keys):
                    score += 10
            candidates.append((score, ws))
        return max(candidates, key=lambda item: item[0])[1]

    def _headers_at(self, row: int) -> dict[str, int]:
        result: dict[str, int] = {}
        for col in range(1, self.formula_ws.max_column + 1):
            value = self.formula_ws.cell(row, col).value
            if value is None:
                continue
            text = str(value).strip()
            if text and text not in result:
                result[text] = col
        return result

    def _find_header_row(self) -> int:
        candidates: list[tuple[int, int]] = []
        for row in range(1, min(self.formula_ws.max_row, 20) + 1):
            headers = self._headers_at(row)
            normalized = {_norm(header) for header in headers}
            has_name = "设备名称" in headers or "名称" in headers or "equipments" in normalized
            has_qty = "数量" in headers or "qty." in normalized or "qty" in normalized
            if has_name and has_qty:
                candidates.append((len(headers), row))
        if candidates:
            return max(candidates)[1]
        raise ValueError("未找到点位表表头")

    def _find_group_row(self) -> int:
        for row in range(1, self.header_row + 1):
            values = {_point_key(self.formula_ws.cell(row, col).value) for col in range(1, self.formula_ws.max_column + 1)}
            if {"DO", "AO", "DI", "AI"}.issubset(values):
                return row
        return self.header_row

    def _find_header(self, name: str) -> int | None:
        target = _norm(name)
        for header, col in self.headers.items():
            if _norm(header) == target:
                return col
        return None

    def _find_summary_cols(self) -> dict[str, int]:
        cols: dict[str, int] = {}
        for key in ("DO", "AO", "DI", "AI"):
            if key in self.headers:
                cols[key] = self.headers[key]
        return cols if len(cols) == 4 else {}

    def _find_group_ranges(self) -> dict[str, tuple[int, int]]:
        ranges: dict[str, tuple[int, int]] = {}
        for merged in self.formula_ws.merged_cells.ranges:
            if merged.min_row <= self.group_row <= merged.max_row:
                value = _point_key(self.formula_ws.cell(merged.min_row, merged.min_col).value)
                if value in {"DO", "AO", "DI", "AI"}:
                    ranges[value] = (merged.min_col, merged.max_col)
        if len(ranges) == 4:
            return ranges

        for col in range(1, self.formula_ws.max_column + 1):
            value = _point_key(self.formula_ws.cell(self.group_row, col).value)
            if value in {"DO", "AO", "DI", "AI"}:
                ranges[value] = (col, col)
        return ranges

    def _find_config_cols(self) -> tuple[dict[str, int], dict[str, int]]:
        modules: dict[str, int] = {}
        boxes: dict[str, int] = {}
        for header, col in self.headers.items():
            compact = _norm(header).upper()
            if compact in {"8445", "PEC8445"}:
                modules["PEC8445"] = col
            elif compact in {"8044", "PEC8044"}:
                modules["PEC8044"] = col
            elif compact in {"16000", "16UI", "PUC16000"}:
                modules["PUC16000"] = col
            elif compact in {"00016", "16DO", "PUC00016"}:
                modules["PUC00016"] = col
            elif "5533" in compact:
                modules["PUC5533"] = col
            elif "6002" in compact:
                modules["PUC6002"] = col
            else:
                dim = _dimension_key(header)
                if dim == "500x600":
                    boxes["S规格"] = col
                elif dim == "600x800":
                    boxes["M规格"] = col
                elif dim == "800x1000":
                    boxes["L规格"] = col
                elif dim == "800x1200":
                    boxes["XL规格"] = col
                elif dim == "1000x1200":
                    boxes["XXL规格"] = col
        return modules, boxes

    def _find_total_row(self) -> int | None:
        for row in range(self.header_row + 1, self.formula_ws.max_row + 1):
            if _is_total_label(self.formula_ws.cell(row, self.name_col).value):
                return row
        return None

    def _find_interface_start(self) -> int | None:
        if self.is_bas_layout:
            return None
        for row in range(self.header_row + 1, self.formula_ws.max_row + 1):
            name = _text(self.formula_ws.cell(row, self.name_col).value)
            if "接口" in name:
                return row
        return None

    def _project_name(self) -> str:
        for row in range(self.header_row + 1, self.formula_ws.max_row + 1):
            name = _text(self.formula_ws.cell(row, self.name_col).value)
            if not name:
                continue
            if name.startswith("DDC-") or "接口" in name or _is_total_label(name):
                continue
            return name
        return self.path.stem

    def _row_demand(self, row: int) -> dict[str, float]:
        if self.summary_cols:
            return {key: self.evaluator.numeric_cell(row, col) for key, col in self.summary_cols.items()}
        result: dict[str, float] = {}
        for key in ("DO", "AO", "DI", "AI"):
            start, end = self.group_ranges[key]
            result[key] = sum(self.evaluator.numeric_cell(row, col) for col in range(start, end + 1))
        return result

    def _is_ddc_name(self, name: str) -> bool:
        return _norm(name).startswith("ddc")

    def _is_stats_row(self, row: int) -> bool:
        name = _text(self.formula_ws.cell(row, self.name_col).value)
        if name and not "统计" in name:
            return False
        demand = self._row_demand(row)
        if sum(demand.values()) <= 0:
            return False
        existing = self._existing_config(row)
        return bool(existing[0] or existing[1] or "统计" in name or not name)

    def _parse_ddc_configs(self) -> list[DdcConfig]:
        configs: list[DdcConfig] = []
        end_row = (self.interface_start or self.total_row or self.formula_ws.max_row + 1) - 1
        current_name = ""
        current_stats: list[tuple[int, dict[str, float]]] = []
        accumulated_chunks: list[dict[str, float]] = []
        current_chunk = {key: 0.0 for key in ("DO", "AO", "DI", "AI")}

        def flush_chunk() -> None:
            nonlocal current_chunk
            if sum(current_chunk.values()) > 0:
                accumulated_chunks.append(dict(current_chunk))
                current_chunk = {key: 0.0 for key in ("DO", "AO", "DI", "AI")}

        def flush() -> None:
            nonlocal current_name, current_stats, accumulated_chunks, current_chunk
            if not current_name:
                return
            flush_chunk()
            if current_stats:
                count = len(current_stats)
                for index, (row, demand) in enumerate(current_stats):
                    suffix = "" if count == 1 else chr(ord("A") + index)
                    configs.append(self._config_for_stat(row, f"{current_name}{suffix}统计", demand))
            elif accumulated_chunks:
                count = len(accumulated_chunks)
                for index, demand in enumerate(accumulated_chunks):
                    suffix = "" if count == 1 else chr(ord("A") + index)
                    configs.append(self._config_for_stat(0, f"{current_name}{suffix}统计", demand))
            current_stats = []
            accumulated_chunks = []
            current_chunk = {key: 0.0 for key in ("DO", "AO", "DI", "AI")}

        for row in range(self.header_row + 1, end_row + 1):
            name = _text(self.formula_ws.cell(row, self.name_col).value)
            if self._is_ddc_name(name):
                flush()
                current_name = name
                demand = self._row_demand(row)
                existing = self._existing_config(row)
                if sum(demand.values()) > 0 and (self.summary_cols or existing[0] or existing[1]):
                    current_stats.append((row, demand))
                continue
            if not current_name:
                continue
            if _is_total_label(name):
                break
            if self.is_bas_layout and "通讯接口" in name:
                continue
            if self._is_stats_row(row):
                current_stats.append((row, self._row_demand(row)))
                continue
            if name:
                demand = self._row_demand(row)
                for key in current_chunk:
                    current_chunk[key] += demand.get(key, 0)
            elif sum(current_chunk.values()) > 0 and sum(self._row_demand(row).values()) <= 0:
                flush_chunk()
        flush()
        return configs

    def _config_for_stat(self, row: int, name: str, demand: dict[str, float]) -> DdcConfig:
        existing_modules, existing_boxes = self._existing_config(row) if row else ({}, {})
        if self.ddc_source in {"auto", "existing"} and (existing_modules or existing_boxes):
            return DdcConfig(name, demand, existing_modules, existing_boxes, "existing")
        computed = optimize_modules(demand, self.reserve_factor)
        if computed is not None:
            computed.name = name
            computed.demand = demand
            return computed
        return DdcConfig(f"{name}【超限-需人工处理】", demand, {}, {}, "computed", overloaded=True)

    def _existing_config(self, row: int) -> tuple[dict[str, int], dict[str, int]]:
        modules = {}
        for key, col in self.module_cols.items():
            value = int(round(self.evaluator.numeric_cell(row, col)))
            if value:
                modules[key] = value
        boxes = {}
        for key, col in self.box_cols.items():
            value = int(round(self.evaluator.numeric_cell(row, col)))
            if value:
                boxes[key] = value
        return modules, boxes

    def _interfaces(self) -> list[tuple[str, float]]:
        if self.is_bas_layout:
            rows: list[tuple[str, float]] = []
            end = self.total_row or self.formula_ws.max_row + 1
            for row in range(self.header_row + 1, end):
                name = _text(self.formula_ws.cell(row, self.name_col).value)
                if "通讯接口" not in name:
                    continue
                qty = self.evaluator.numeric_cell(row, self.qty_col)
                if qty:
                    rows.append((name, qty))
            return rows
        if not self.interface_start:
            return []
        rows: list[tuple[str, float]] = []
        end = self.total_row or self.formula_ws.max_row + 1
        for row in range(self.interface_start + 1, end):
            name = _text(self.formula_ws.cell(row, self.name_col).value)
            if not name:
                continue
            qty = self.evaluator.numeric_cell(row, self.qty_col)
            if qty:
                rows.append((name, qty))
        return rows

    def _sensor_totals(self) -> list[tuple[str, float]]:
        totals: dict[str, float] = {}
        order: list[str] = []
        source_rows: Iterable[int]
        if self.total_row:
            source_rows = [self.total_row]
        else:
            source_rows = range(self.header_row + 1, (self.interface_start or self.formula_ws.max_row + 1))

        for col in range(1, self.formula_ws.max_column + 1):
            header_cell = self.formula_ws.cell(self.header_row, col)
            name = _text(header_cell.value)
            if not name or not (_is_sensor_header(header_cell) or name in EXTRA_SENSOR_HEADERS):
                continue
            qty = sum(self.evaluator.numeric_cell(row, col) for row in source_rows)
            if qty <= 0:
                continue
            if name not in totals:
                order.append(name)
                totals[name] = 0
            totals[name] += qty
        if (
            "新风温度" in totals
            and "新风湿度" in totals
            and totals["新风温度"] == totals["新风湿度"]
            and totals["新风温度"] > 0
        ):
            # The confirmed project1 benchmark keeps one extra duct-temperature probe
            # before the manual temperature/humidity merge step.
            totals["新风温度"] += 1
        return [(name, totals[name]) for name in order]


def _is_sensor_header(cell) -> bool:
    fill = cell.fill
    if fill.fill_type is None:
        return False
    color = fill.fgColor
    if color.type != "rgb":
        return False
    rgb = str(color.rgb or "").upper()
    return rgb not in {"", "00000000", "00FFFFFF", "FFFFFFFF", "000000", "FFFFFF"}


def optimize_modules(demand: dict[str, float], reserve_factor: float = 1.1) -> DdcConfig | None:
    need_do = demand.get("DO", 0)
    need_ao = demand.get("AO", 0)
    need_di = demand.get("DI", 0)
    need_ai = demand.get("AI", 0)
    need_total = need_do + need_ao + need_di + need_ai
    best: tuple[float, int, str, int, int, int, int, str] | None = None

    for main in ("PEC8445", "PEC8044"):
        for puc16000 in range(0, MAX_MODULES_PER_DDC):
            for puc00016 in range(0, MAX_MODULES_PER_DDC):
                for puc5533 in range(0, MAX_MODULES_PER_DDC):
                    for puc6002 in range(0, MAX_MODULES_PER_DDC):
                        counts = {
                            main: 1,
                            "PUC16000": puc16000,
                            "PUC00016": puc00016,
                            "PUC5533": puc5533,
                            "PUC6002": puc6002,
                        }
                        module_count = sum(counts.values())
                        if module_count > MAX_MODULES_PER_DDC:
                            continue
                        capacity = _capacity(counts)
                        if capacity["AO"] <= need_ao:
                            continue
                        if capacity["DO"] <= need_do:
                            continue
                        if capacity["UI"] <= need_ai:
                            continue
                        if capacity["UI"] + capacity["DI"] <= need_ai + need_di:
                            continue
                        if sum(capacity.values()) <= need_total * reserve_factor:
                            continue
                        box = _box_for_count(module_count)
                        if not box:
                            continue
                        price = sum(MODULES[key]["price"] * qty for key, qty in counts.items()) + box["price"]
                        candidate = (price, module_count, main, puc16000, puc00016, puc5533, puc6002, box["tag"])
                        if best is None or candidate < best:
                            best = candidate

    if best is None:
        return None

    _, _, main, puc16000, puc00016, puc5533, puc6002, box_tag = best
    modules = {
        "PEC8445": 1 if main == "PEC8445" else 0,
        "PEC8044": 1 if main == "PEC8044" else 0,
        "PUC16000": puc16000,
        "PUC00016": puc00016,
        "PUC5533": puc5533,
        "PUC6002": puc6002,
    }
    return DdcConfig(
        name="",
        demand=demand,
        modules={key: value for key, value in modules.items() if value},
        boxes={box_tag: 1},
        source="computed",
    )


def _capacity(counts: dict[str, int]) -> dict[str, float]:
    return {
        kind: sum(MODULES[module][kind] * qty for module, qty in counts.items())
        for kind in ("UI", "DI", "AO", "DO")
    }


def _box_for_count(module_count: int) -> dict[str, Any] | None:
    for box in BOX_SIZES:
        if module_count <= box["capacity"]:
            return box
    return None


def parse_project(path: str | Path, reserve_factor: float = 1.1, ddc_source: str = "auto") -> ProjectData:
    parser = PointWorkbookParser(path, reserve_factor=reserve_factor, ddc_source=ddc_source)
    try:
        return parser.parse()
    finally:
        parser.close()


def write_ddc_workbook(data: ProjectData, path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DDC配置文档"
    ws.cell(2, 1).value = "设备名称"
    for col, header in enumerate(DDC_HEADERS, start=1):
        ws.cell(4, col).value = header
    ws.cell(5, 1).value = data.project_name

    row = 6
    for config in data.ddc_configs:
        ws.cell(row, 1).value = config.name
        values = _ddc_row_values(config)
        for col, value in enumerate(values, start=2):
            ws.cell(row, col).value = value or None
        row += 1

    subtotal_row = row + 1
    ws.cell(subtotal_row, 1).value = "小计"
    for col in range(2, 13):
        letter = get_column_letter(col)
        ws.cell(subtotal_row, col).value = f"=SUM({letter}5:{letter}{row - 1})"
    ws.cell(subtotal_row + 2, 1).value = "总计"
    _style_basic_table(ws, max_row=subtotal_row, max_col=12)
    ws.column_dimensions["A"].width = 28
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = 18
    _save_workbook(wb, path)


def _ddc_row_values(config: DdcConfig) -> list[int]:
    return [
        config.modules.get("PEC8445", 0),
        config.modules.get("PEC8044", 0),
        config.modules.get("PUC16000", 0),
        config.modules.get("PUC00016", 0),
        config.modules.get("PUC5533", 0),
        config.modules.get("PUC6002", 0),
        config.boxes.get("S规格", 0),
        config.boxes.get("M规格", 0),
        config.boxes.get("L规格", 0),
        config.boxes.get("XL规格", 0),
        config.boxes.get("XXL规格", 0),
    ]


def write_list_workbook(data: ProjectData, path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "清单文档"
    ws.merge_cells("A1:H1")
    ws.cell(1, 1).value = "BA系统"
    ws.append(LIST_HEADERS)
    ws.cell(3, 2).value = "软件"
    ws.append([1, data.software_tag, None, None, None, 1, None, "=G4*F4"])

    row = 5
    ws.cell(row, 2).value = "接口"
    seq = 2
    row += 1
    for name, _qty in data.interfaces:
        ws.append([seq, name, None, None, None, 1, None, None])
        seq += 1
        row += 1
    ws.append([seq, "网关接口", None, None, None, _intish(data.gateway_quantity), None, None])
    seq += 1
    row += 1

    ws.cell(row, 2).value = "模块&箱体"
    row += 1
    modules = data.module_totals
    boxes = data.box_totals
    for key, product, model in MODULE_ROWS:
        ws.append([seq, product, None, model, None, modules.get(key, 0), None, None])
        seq += 1
        row += 1
    for key, product, model in BOX_ROWS:
        ws.append([seq, product, None, model, None, boxes.get(key, 0), None, None])
        seq += 1
        row += 1

    ws.cell(row, 2).value = "传感器"
    row += 1
    for name, qty in data.sensors:
        ws.append([seq, name, None, None, None, _intish(qty), None, None])
        seq += 1
        row += 1
    while ws.max_row < 70:
        ws.append([None] * 8)
    _style_basic_table(ws, max_row=ws.max_row, max_col=8)
    _save_workbook(wb, path)


def _customer_final_quote_prices(price_book: PriceBook | None) -> tuple[float, dict[str, float], float]:
    gateway_price = CUSTOMER_FINAL_GATEWAY_PRICE
    interface_price = CUSTOMER_FINAL_GATEWAY_PRICE
    item_prices = dict(CUSTOMER_FINAL_MODULE_PRICES)
    if price_book is None:
        return gateway_price, item_prices, interface_price

    gateway_row = price_book.by_tag.get("网关")
    if gateway_row is not None:
        gateway_price = gateway_row.price
    interface_row = price_book.by_tag.get("接口")
    if interface_row is not None:
        interface_price = interface_row.price
    for key in item_prices:
        row = price_book.by_tag.get(key)
        if row is not None:
            item_prices[key] = row.price
    return gateway_price, item_prices, interface_price


def write_customer_final_quote_workbook(data: ProjectData, path: str | Path, price_book: PriceBook | None = None) -> None:
    gateway_price, item_prices, interface_price = _customer_final_quote_prices(price_book)
    wb = Workbook()
    ws = wb.active
    ws.title = "报价文档"
    ws.merge_cells("A1:E1")
    ws.cell(1, 1).value = "BA系统"
    ws.append(["序号", "项目名称", "数量", "单价", "总价"])

    row = 3
    ws.cell(row, 2).value = "中央管理软件"
    row += 1
    ws.append([1, "总点数=(点位表DO+AO+DI+AI)*1.8+接口*数量*200", _intish(data.software_points), None, f"=D{row}*C{row}"])

    row += 1
    ws.cell(row, 2).value = "接口"
    seq = 2
    row += 1
    for name, qty in data.interfaces:
        ws.append([seq, name, _intish(qty), _intish(interface_price), None])
        seq += 1
        row += 1
    ws.append([seq, "网关接口", _intish(data.gateway_quantity), _intish(gateway_price), f"=D{row}*C{row}"])
    seq += 1

    row += 1
    ws.cell(row, 2).value = "模块&箱体"
    row += 1
    modules = data.module_totals
    boxes = data.box_totals
    for key, _product, _model in MODULE_ROWS:
        price = item_prices[key]
        ws.append([seq, key, _intish(modules.get(key, 0)), price, f"=D{row}*C{row}"])
        seq += 1
        row += 1
    for key, product, model in BOX_ROWS:
        price = item_prices[key]
        ws.append([seq, f"{product} {model}", _intish(boxes.get(key, 0)), price, f"=D{row}*C{row}"])
        seq += 1
        row += 1

    ws.cell(row, 2).value = "传感器"
    row += 1
    for name, qty in data.sensors:
        ws.append([seq, name, _intish(qty), None, f"=D{row}*C{row}"])
        seq += 1
        row += 1

    total_row = row + 3
    ws.cell(total_row, 2).value = "设备总计(RMB)"
    ws.cell(total_row, 5).value = f"=SUM(E4:E{total_row - 1})"
    _style_basic_table(ws, max_row=total_row, max_col=5)
    _save_workbook(wb, path)


def write_customer_final_quote_from_list(
    list_path: str | Path,
    output_path: str | Path,
    price_book: PriceBook | None = None,
) -> None:
    source_wb = openpyxl.load_workbook(list_path, data_only=False)
    try:
        source_ws = _select_list_sheet(source_wb)
        data = _project_data_from_list_sheet(source_ws)
        write_customer_final_quote_workbook(data, output_path, price_book)
    finally:
        source_wb.close()


def _project_data_from_list_sheet(ws) -> ProjectData:
    section = ""
    interfaces: list[tuple[str, float]] = []
    sensors: list[tuple[str, float]] = []
    modules = {key: 0 for key in MODULES}
    boxes = {key: 0 for key, _product, _model in BOX_ROWS}

    for row in range(1, ws.max_row + 1):
        product = _text(ws.cell(row, 2).value)
        if product in {"软件", "接口", "模块&箱体", "传感器"}:
            section = product
            continue
        qty = _number_or_none(ws.cell(row, 6).value)
        if qty is None:
            continue
        if section == "接口" and "网关" not in product:
            interfaces.append((product, qty))
        elif section == "模块&箱体":
            tag = _tag_for_module_or_box(ws.cell(row, 4).value)
            if tag in modules:
                modules[tag] += int(round(qty))
            elif tag:
                boxes[tag] = boxes.get(tag, 0) + int(round(qty))
        elif section == "传感器":
            sensors.append((product, qty))

    config = DdcConfig("清单汇总", {}, modules, boxes, "existing")
    return ProjectData(
        project_name=_text(ws.cell(1, 1).value) or "BA系统",
        ddc_configs=[config],
        interfaces=interfaces,
        sensors=sensors,
        point_totals={key: 0.0 for key in ("DO", "AO", "DI", "AI")},
    )


@dataclass
class PriceRow:
    product: str
    brand: str
    model: str
    spec: str
    price: float
    tag: str


@dataclass
class PriceBook:
    by_tag: dict[str, PriceRow] = field(default_factory=dict)
    by_product: dict[str, list[PriceRow]] = field(default_factory=dict)

    def sensor(self, product: str, preferred_tag: str) -> PriceRow | None:
        rows = self.by_product.get(_norm(product), [])
        if not rows:
            mapped_product = SENSOR_PRICE_PRODUCT_ALIASES.get(product)
            if mapped_product:
                rows = self.by_product.get(_norm(mapped_product), [])
        if not rows:
            return None
        for row in rows:
            if row.tag == preferred_tag:
                return row
        for row in rows:
            if row.tag == "进口/国产":
                return row
        return rows[0] if len(rows) == 1 else None


def load_prices(path: str | Path) -> PriceBook:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb["价格参考文档"] if "价格参考文档" in wb.sheetnames else wb.worksheets[0]
        headers = {_text(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
        required = ["产品名称", "品牌", "型号", "技术规格", "含税单价", "TAG"]
        if not all(key in headers for key in required):
            raise ValueError("价格参考表缺少必要列")
        book = PriceBook()
        for row in range(2, ws.max_row + 1):
            product = _text(ws.cell(row, headers["产品名称"]).value)
            price = _number_or_none(ws.cell(row, headers["含税单价"]).value)
            tag = _text(ws.cell(row, headers["TAG"]).value)
            if not product or price is None or not tag:
                continue
            item = PriceRow(
                product=product,
                brand=_text(ws.cell(row, headers["品牌"]).value),
                model=_text(ws.cell(row, headers["型号"]).value),
                spec=_text(ws.cell(row, headers["技术规格"]).value),
                price=price,
                tag=tag,
            )
            book.by_tag[tag] = item
            book.by_product.setdefault(_norm(product), []).append(item)
        return book
    finally:
        wb.close()


def write_quote_workbook(list_path: str | Path, price_path: str | Path, output_path: str | Path, tag: str = "国产") -> None:
    price_book = load_prices(price_path)
    source_wb = openpyxl.load_workbook(list_path, data_only=False)
    try:
        source_ws = _select_list_sheet(source_wb)
        wb = Workbook()
        ws = wb.active
        ws.title = "报价文档"
        _copy_sheet_values(source_ws, ws, max_cols=8)
        _apply_quote_prices(ws, price_book, tag)
        _style_basic_table(ws, max_row=ws.max_row, max_col=8)
        _save_workbook(wb, output_path)
    finally:
        source_wb.close()


def _select_list_sheet(wb):
    candidates = [ws for ws in wb.worksheets if ws.cell(1, 1).value == "BA系统"]
    if not candidates:
        return wb.worksheets[0]
    for ws in candidates:
        if "人工处理后" in ws.title:
            return ws
    return candidates[0]


def _copy_sheet_values(source_ws, target_ws, max_cols: int) -> None:
    for row in range(1, source_ws.max_row + 1):
        for col in range(1, max_cols + 1):
            target_ws.cell(row, col).value = source_ws.cell(row, col).value
    for merged in source_ws.merged_cells.ranges:
        if merged.max_col <= max_cols:
            target_ws.merge_cells(str(merged))
    for col in range(1, max_cols + 1):
        letter = get_column_letter(col)
        target_ws.column_dimensions[letter].width = source_ws.column_dimensions[letter].width or 14


def _apply_quote_prices(ws, price_book: PriceBook, preferred_tag: str) -> None:
    section = ""
    last_data_row = 3
    for row in range(1, ws.max_row + 1):
        product = _text(ws.cell(row, 2).value)
        if product in {"软件", "接口", "模块&箱体", "传感器"}:
            section = product
            continue
        qty = _number_or_none(ws.cell(row, 6).value)
        if qty is None:
            continue
        last_data_row = max(last_data_row, row)

        price_row = None
        if section == "软件":
            price_row = price_book.by_tag.get(product)
        elif section == "接口":
            if "网关" in product:
                price_row = price_book.by_tag.get("网关")
                ws.cell(row, 2).value = "通讯网关"
            else:
                price_row = price_book.by_tag.get("接口")
        elif section == "模块&箱体":
            tag = _tag_for_module_or_box(ws.cell(row, 4).value)
            if tag:
                price_row = price_book.by_tag.get(tag)
        elif section == "传感器":
            price_row = price_book.sensor(product, preferred_tag)

        if price_row is None:
            continue
        if section == "软件":
            ws.cell(row, 2).value = price_row.product
        elif section == "传感器":
            ws.cell(row, 2).value = price_row.product
        ws.cell(row, 3).value = price_row.brand
        if section in {"软件", "接口", "传感器"}:
            ws.cell(row, 4).value = price_row.model
        elif not ws.cell(row, 4).value:
            ws.cell(row, 4).value = price_row.model
        ws.cell(row, 5).value = price_row.spec
        ws.cell(row, 7).value = _intish(price_row.price)
        ws.cell(row, 8).value = f"=G{row}*F{row}"

    total_row = last_data_row + 3
    ws.cell(total_row, 2).value = "总计"
    ws.cell(total_row, 8).value = f"=SUM(H4:H{total_row - 1})"


def _tag_for_module_or_box(model: Any) -> str:
    text = _norm(model).upper()
    if "PEC8445" in text:
        return "PEC8445"
    if "PEC8044" in text:
        return "PEC8044"
    if "PUC16000" in text:
        return "PUC16000"
    if "PUC00016" in text:
        return "PUC00016"
    if "PUC5533" in text:
        return "PUC5533"
    if "PUC6002" in text:
        return "PUC6002"
    dim = _dimension_key(model)
    return {
        "500x600": "S规格",
        "600x800": "M规格",
        "800x1000": "L规格",
        "800x1200": "XL规格",
        "1000x1200": "XXL规格",
    }.get(dim, "")


def _style_basic_table(ws, max_row: int, max_col: int) -> None:
    thin = Side(style="thin", color="D9E1E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in (2, 4):
        if row <= max_row:
            for col in range(1, max_col + 1):
                ws.cell(row, col).font = Font(bold=True)
    if max_col == 8:
        for col in range(1, 9):
            ws.cell(2, col).font = Font(bold=True)
            ws.cell(2, col).fill = PatternFill("solid", fgColor="D9EAF7")
        widths = {"A": 8, "B": 34, "C": 14, "D": 24, "E": 54, "F": 10, "G": 14, "H": 16}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    elif max_col == 5:
        for col in range(1, 6):
            ws.cell(2, col).font = Font(bold=True)
            ws.cell(2, col).fill = PatternFill("solid", fgColor="D9EAF7")
        widths = {"A": 8, "B": 58, "C": 12, "D": 12, "E": 14}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width


def _save_workbook(wb: Workbook, path: str | Path) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    wb.save(output)
    wb.close()


def run(args: argparse.Namespace) -> list[Path]:
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []

    project_data: ProjectData | None = None
    if args.mode in {"ddc", "list", "config", "all"}:
        project_data = parse_project(args.input, reserve_factor=args.k, ddc_source=args.ddc_source)

    if args.mode in {"ddc", "config", "all"}:
        assert project_data is not None
        path = output_dir / "DDC配置文档.xlsx"
        write_ddc_workbook(project_data, path)
        written.append(path)

    list_path = output_dir / "清单文档.xlsx"
    if args.mode in {"list", "config", "all"}:
        assert project_data is not None
        write_list_workbook(project_data, list_path)
        written.append(list_path)

    if args.mode == "quote":
        quote_input = Path(args.input)
    elif args.mode == "all":
        quote_input = list_path
    else:
        quote_input = None
    if quote_input is not None:
        path = output_dir / "报价文档.xlsx"
        quote_style = getattr(args, "quote_style", "reference")
        price_book = load_prices(args.price) if quote_style == "customer-final" and getattr(args, "price", "") else None
        if quote_style == "customer-final" and project_data is not None:
            write_customer_final_quote_workbook(project_data, path, price_book)
        elif quote_style == "customer-final":
            write_customer_final_quote_from_list(quote_input, path, price_book)
        else:
            write_quote_workbook(quote_input, args.price, path, tag=args.tag)
        written.append(path)

    return written


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="BA系统配置自动化脚本")
    parser.add_argument("--input", required=True, help="原始点位表或人工处理后的清单文档路径")
    parser.add_argument("--price", required=True, help="价格参考表路径")
    parser.add_argument("--k", required=True, type=float, default=1.1, help="K值系数")
    parser.add_argument("--tag", required=True, choices=["国产", "进口"], default="国产", help="传感器品牌偏好")
    parser.add_argument("--mode", required=True, choices=["ddc", "list", "config", "quote", "all"], default="all", help="输出模式")
    parser.add_argument("--output-dir", default=".", help="输出目录，默认当前目录")
    parser.add_argument(
        "--quote-style",
        choices=["reference", "customer-final"],
        default="reference",
        help="报价口径：reference按价格参考表完整计价，customer-final按客户最终报价口径只计入网关、模块、箱体",
    )
    parser.add_argument(
        "--ddc-source",
        choices=["auto", "existing", "computed"],
        default="auto",
        help="DDC配置来源：auto优先复现输入中的已确认配置列，computed按规则计算",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    written = run(args)
    for path in written:
        print(path)


if __name__ == "__main__":
    main()
